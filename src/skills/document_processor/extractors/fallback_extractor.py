import os
from typing import List

from src.skills.document_processor.extractors.base import BaseExtractor, ExtractedDocument
from src.logger import get_logger

logger = get_logger(__name__)

class FallbackExtractor(BaseExtractor):
    """Extrator de fallback para PDFs (via pdfplumber) e XLSX (via openpyxl)."""

    @property
    def supported_extensions(self) -> List[str]:
        return [".pdf", ".xlsx"]

    def extract(self, file_path: str) -> ExtractedDocument:
        filename = os.path.basename(file_path)
        format_str = file_path.lower().split('.')[-1]

        if format_str == "pdf":
            return self._extract_pdf(file_path, filename)
        elif format_str == "xlsx":
            return self._extract_xlsx(file_path, filename)
        
        return ExtractedDocument(
            source_path=file_path,
            format=format_str,
            title=filename,
            text_content="",
            extraction_errors=["Formato não suportado pelo fallback."],
        )

    def _extract_pdf(self, file_path: str, filename: str) -> ExtractedDocument:
        try:
            import pdfplumber
            text_content = ""
            num_pages = 0
            with pdfplumber.open(file_path) as pdf:
                num_pages = len(pdf.pages)
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_content += page_text + "\n\n"
            
            return ExtractedDocument(
                source_path=file_path,
                format="pdf",
                title=filename,
                text_content=text_content.strip(),
                metadata={"extractor": "FallbackExtractor", "tool": "pdfplumber"},
                num_pages=num_pages,
            )
        except ImportError:
            return ExtractedDocument(
                source_path=file_path,
                format="pdf",
                title=filename,
                text_content="",
                extraction_errors=["pdfplumber não está instalado."],
            )
        except Exception as e:
            logger.error(f"Erro no pdfplumber para {file_path}: {e}")
            return ExtractedDocument(
                source_path=file_path,
                format="pdf",
                title=filename,
                text_content="",
                extraction_errors=[str(e)],
            )

    def _extract_xlsx(self, file_path: str, filename: str) -> ExtractedDocument:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text_lines = []
            tables = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_lines.append(f"## Planilha: {sheet_name}")
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    row_str = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_str):  # Evitar linhas vazias
                        text_lines.append(",".join(row_str))
                        sheet_data.append(row_str)
                
                if sheet_data:
                    tables.append({"sheet": sheet_name, "header": sheet_data[0], "data": sheet_data[1:]})
                text_lines.append("\n")

            return ExtractedDocument(
                source_path=file_path,
                format="xlsx",
                title=filename,
                text_content="\n".join(text_lines),
                metadata={"extractor": "FallbackExtractor", "tool": "openpyxl"},
                tables=tables,
            )
        except ImportError:
            return ExtractedDocument(
                source_path=file_path,
                format="xlsx",
                title=filename,
                text_content="",
                extraction_errors=["openpyxl não está instalado."],
            )
        except Exception as e:
            logger.error(f"Erro no openpyxl para {file_path}: {e}")
            return ExtractedDocument(
                source_path=file_path,
                format="xlsx",
                title=filename,
                text_content="",
                extraction_errors=[str(e)],
            )
